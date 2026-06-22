"""
routes/admin.py — Admin Blueprint  (README v3 — patched + Excel→users + bulk manage)

แก้ไข/เพิ่มจากเดิม:
  - create_user()           — ส่งต่อ role ที่เลือกจากฟอร์ม
  - manage_voters()/export_voters() — รายชื่อผู้มาใช้สิทธิ
  - import_members()        — นำเข้า Excel เป็น users (อ่าน A=ชื่อ B=username C=email D=password)
  - imported_users()        — หน้าจัดการผู้ใช้ที่นำเข้าจาก Excel (แก้/ลบเป็นชุด)
  - bulk_delete_users()     — ลบผู้ใช้ที่เลือกหลายคนพร้อมกัน
  - update_user_profile()   — แก้ไขชื่อ/username/email/รหัสผ่าน ของผู้ใช้รายคน
"""

import io
import functools
import os
import uuid
from datetime import datetime

from flask import (
    Blueprint, render_template, redirect, url_for,
    flash, request, send_file, abort, current_app,
)
from flask_login import login_required, current_user

from models.election       import Election
from models.candidate      import Candidate
from models.vote           import Vote, Turnout
from models.member         import Member
from models.system_setting import SystemSetting
from models.access_log     import AccessLog
from routes.export_utils   import make_table_response

admin_bp = Blueprint("admin", __name__)


# ── Guard ──────────────────────────────────────────────────

def admin_required(f):
    @functools.wraps(f)
    @login_required
    def wrapper(*args, **kwargs):
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return wrapper


# ── Helper: xlsx ───────────────────────────────────────────

def _make_xlsx(ws_title: str, headers: list, rows: list, col_widths: list) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ws_title

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Helper: บันทึกรูปผู้สมัคร ───────────────────────────────

ALLOWED_IMG_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp"}

def _save_candidate_photo(file_storage):
    """บันทึกไฟล์รูปผู้สมัคร -> คืน URL path (/static/...) หรือ None ถ้าไม่มีไฟล์
    raise ValueError ถ้านามสกุลไฟล์ไม่รองรับ"""
    if not file_storage or not file_storage.filename:
        return None
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in ALLOWED_IMG_EXT:
        raise ValueError("รองรับเฉพาะไฟล์รูป .jpg .jpeg .png .gif .webp")
    upload_dir = os.path.join(current_app.static_folder, "uploads", "candidates")
    os.makedirs(upload_dir, exist_ok=True)
    fname = f"{uuid.uuid4().hex}{ext}"
    file_storage.save(os.path.join(upload_dir, fname))
    return f"/static/uploads/candidates/{fname}"

# ── Dashboard ──────────────────────────────────────────────

@admin_bp.route("/")
@admin_required
def dashboard():
    elections = Election.get_all()
    stats = []
    for e in elections:
        candidates  = Candidate.get_by_election_with_votes(e.id)
        total_votes = sum(c.vote_count for c in candidates)
        stats.append({
            "election":    e,
            "candidates":  len(candidates),
            "total_votes": total_votes,
        })
    total_members  = len(Member.get_all())
    total_verified = Member.count_verified()
    return render_template(
        "admin/dashboard.html",
        stats=stats,
        total_members=total_members,
        total_verified=total_verified,
    )


# ── Elections ──────────────────────────────────────────────

@admin_bp.route("/elections")
@admin_required
def elections():
    all_elections = Election.get_all()
    return render_template("admin/elections.html", elections=all_elections)


@admin_bp.route("/elections/create", methods=["POST"])
@admin_required
def create_election():
    title         = request.form.get("title", "").strip()
    election_type = request.form.get("type", "committee")
    max_votes     = request.form.get("max_votes", 1, type=int)
    is_visible    = bool(request.form.get("is_visible"))

    if not title:
        flash("กรุณาระบุชื่อวาระ", "danger")
        return redirect(url_for("admin.elections"))

    Election.create(
        title, election_type, max_votes=max_votes,
        is_visible=is_visible, created_by=current_user.id,
    )
    flash(f"สร้างวาระ '{title}' สำเร็จ", "success")
    return redirect(url_for("admin.elections"))


@admin_bp.route("/elections/<int:election_id>/status", methods=["POST"])
@admin_required
def set_election_status(election_id: int):
    election = Election.get_by_id(election_id)
    if not election:
        abort(404)
    status = request.form.get("status")
    try:
        election.set_status(status)
        flash(f"เปลี่ยนสถานะวาระเป็น '{status}' แล้ว", "success")
    except ValueError as e:
        flash(str(e), "danger")
    return redirect(url_for("admin.elections"))


@admin_bp.route("/elections/<int:election_id>/settings", methods=["POST"])
@admin_required
def update_election_settings(election_id: int):
    election = Election.get_by_id(election_id)
    if not election:
        abort(404)
    title      = request.form.get("title", "").strip() or None
    max_votes  = request.form.get("max_votes", type=int)
    is_visible = request.form.get("is_visible")
    is_visible = True if is_visible == "1" else False if is_visible == "0" else None
    election.update_settings(title=title, max_votes=max_votes, is_visible=is_visible)
    flash("บันทึกการตั้งค่าวาระแล้ว", "success")
    return redirect(url_for("admin.elections"))


@admin_bp.route("/elections/<int:election_id>/delete", methods=["POST"])
@admin_required
def delete_election(election_id: int):
    election = Election.get_by_id(election_id)
    if not election:
        abort(404)
    election.delete()
    flash("ลบวาระแล้ว", "info")
    return redirect(url_for("admin.elections"))


# ── Candidates ─────────────────────────────────────────────

@admin_bp.route("/elections/<int:election_id>/candidates")
@admin_required
def manage_candidates(election_id: int):
    election = Election.get_by_id(election_id)
    if not election:
        abort(404)
    candidates = Candidate.get_by_election_with_votes(election_id)
    return render_template(
        "admin/candidates.html", election=election, candidates=candidates
    )


@admin_bp.route("/elections/<int:election_id>/candidates/add", methods=["POST"])
@admin_required
def add_candidate(election_id: int):
    election = Election.get_by_id(election_id)
    if not election:
        abort(404)
    name      = request.form.get("name", "").strip()
    party     = request.form.get("party", "").strip()
    bio       = request.form.get("bio", "").strip()
    number    = request.form.get("number", type=int)
    photo_url = request.form.get("photo_url", "").strip()

    try:
        uploaded = _save_candidate_photo(request.files.get("photo"))
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("admin.manage_candidates", election_id=election_id))
    if uploaded:
        photo_url = uploaded

    if not name:
        flash("กรุณาระบุชื่อผู้สมัคร", "danger")
        return redirect(url_for("admin.manage_candidates", election_id=election_id))
    if not number:
        flash("กรุณาระบุหมายเลขผู้สมัคร", "danger")
        return redirect(url_for("admin.manage_candidates", election_id=election_id))

    Candidate.create(
        election_id, name,
        party=party, bio=bio, photo_url=photo_url, number=number,
    )
    flash(f"เพิ่มผู้สมัคร '{name}' สำเร็จ", "success")
    return redirect(url_for("admin.manage_candidates", election_id=election_id))


@admin_bp.route("/candidates/<int:candidate_id>/edit", methods=["POST"])
@admin_required
def edit_candidate(candidate_id: int):
    candidate = Candidate.get_by_id(candidate_id)
    if not candidate:
        abort(404)
    name      = request.form.get("name", "").strip()
    party     = request.form.get("party", "").strip()
    bio       = request.form.get("bio", "").strip()
    number    = request.form.get("number", type=int)
    photo_url_text = request.form.get("photo_url", "").strip()

    if not name:
        flash("กรุณาระบุชื่อผู้สมัคร", "danger")
        return redirect(url_for("admin.manage_candidates", election_id=candidate.election_id))

    try:
        uploaded = _save_candidate_photo(request.files.get("photo"))
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("admin.manage_candidates", election_id=candidate.election_id))

    if uploaded:
        new_photo = uploaded
    elif photo_url_text:
        new_photo = photo_url_text
    else:
        new_photo = None

    candidate.update(
        name, party=party, bio=bio, photo_url=new_photo, number=number,
    )
    flash("แก้ไขข้อมูลผู้สมัครแล้ว", "success")
    return redirect(url_for("admin.manage_candidates", election_id=candidate.election_id))


@admin_bp.route("/candidates/<int:candidate_id>/delete", methods=["POST"])
@admin_required
def delete_candidate(candidate_id: int):
    candidate = Candidate.get_by_id(candidate_id)
    if not candidate:
        abort(404)
    election_id = candidate.election_id
    candidate.delete()
    flash("ลบผู้สมัครแล้ว", "info")
    return redirect(url_for("admin.manage_candidates", election_id=election_id))


# ── Export results Excel ───────────────────────────────────

@admin_bp.route("/elections/<int:election_id>/results/export")
@admin_required
def export_results(election_id: int):
    election   = Election.get_by_id(election_id)
    candidates = Candidate.get_by_election_with_votes(election_id)
    total      = sum(c.vote_count for c in candidates)
    rows = [
        (c.number, c.name, c.vote_count,
         f"{round(c.vote_count / total * 100, 1)}%" if total else "0%")
        for c in candidates
    ]
    buf = _make_xlsx(
        "ผลคะแนน",
        ["หมายเลข", "ชื่อ", "คะแนน", "เปอร์เซ็นต์"],
        rows,
        [10, 30, 12, 12],
    )
    filename = f"results_{election_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=filename)


# ── Voters (ผู้มาใช้สิทธิ — ตรวจสอบได้) ─────────────────────

@admin_bp.route("/elections/<int:election_id>/voters")
@admin_required
def manage_voters(election_id: int):
    """รายชื่อผู้มาใช้สิทธิในวาระ — ดึงจาก Turnout (member จริง)"""
    election = Election.get_by_id(election_id)
    if not election:
        abort(404)
    voters = Turnout.list_by_election(election_id)
    return render_template("admin/voters.html", election=election, voters=voters)


@admin_bp.route("/elections/<int:election_id>/voters/export")
@admin_required
def export_voters(election_id: int):
    """export รายชื่อผู้มาใช้สิทธิ (Excel/PDF)"""
    fmt = request.args.get("format", "excel")
    election = Election.get_by_id(election_id)
    if not election:
        abort(404)
    voters = Turnout.list_by_election(election_id)
    rows = [
        (i + 1, v.get("full_name") or f"#{v['member_id']}", str(v.get("voted_at", "")))
        for i, v in enumerate(voters)
    ]
    return make_table_response(
        fmt,
        sheet_title="ผู้มาใช้สิทธิ",
        headers=["ลำดับ", "ชื่อ-สกุล", "เวลาที่ลงคะแนน"],
        rows=rows, col_widths=[8, 30, 22],
        filename_base=f"voters_{election_id}",
    )


# ── Users Management ───────────────────────────────────────

@admin_bp.route("/users")
@admin_required
def manage_users():
    from models.user import User
    users = User.get_all()
    return render_template("admin/users.html", users=users)


@admin_bp.route("/users/create", methods=["POST"])
@admin_required
def create_user():
    from models.user import User
    username  = request.form.get("username", "").strip()
    email     = request.form.get("email", "").strip().lower()
    full_name = request.form.get("full_name", "").strip()
    password  = request.form.get("password", "")
    role      = request.form.get("role", "voter")
    if not all([username, email, full_name, password]):
        flash("กรุณากรอกข้อมูลให้ครบ", "danger")
        return redirect(url_for("admin.manage_users"))

    user = User.create(username, email, password, full_name)
    if role == "admin":
        try:
            user.set_role("admin")
        except ValueError:
            pass
    flash(f"เพิ่มผู้ใช้งาน '{username}' ({role}) แล้ว", "success")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/<int:user_id>/edit", methods=["POST"])
@admin_required
def edit_user(user_id: int):
    from models.user import User
    user = User.get_by_id_any(user_id)
    if not user:
        abort(404)
    password = request.form.get("password", "").strip()
    if password:
        user.set_password(password)
        flash("เปลี่ยนรหัสผ่านแล้ว", "success")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def delete_user(user_id: int):
    from models.user import User
    user = User.get_by_id_any(user_id)
    if not user or user.id == current_user.id:
        flash("ไม่สามารถลบได้", "warning")
        return redirect(url_for("admin.manage_users"))
    user.delete()
    flash("ลบผู้ใช้งานแล้ว", "info")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/<int:user_id>/role", methods=["POST"])
@admin_required
def set_user_role(user_id: int):
    from models.user import User
    user = User.get_by_id_any(user_id)
    if not user:
        abort(404)
    if user.id == current_user.id:
        flash("ไม่สามารถเปลี่ยนโรลของตัวเองได้", "warning")
        return redirect(url_for("admin.manage_users"))
    new_role = request.form.get("role")
    try:
        user.set_role(new_role)
        flash(f"เปลี่ยน '{user.username}' เป็น {new_role} แล้ว", "success")
    except ValueError as e:
        flash(str(e), "danger")
    return redirect(url_for("admin.manage_users"))


@admin_bp.route("/users/<int:user_id>/toggle-active", methods=["POST"])
@admin_required
def toggle_user_active(user_id: int):
    from models.user import User
    user = User.get_by_id_any(user_id)
    if not user:
        abort(404)
    if user.id == current_user.id:
        flash("ไม่สามารถระงับบัญชีของตัวเองได้", "warning")
        return redirect(url_for("admin.manage_users"))
    user.set_active(not user.is_active)
    status = "เปิดใช้" if user.is_active else "ระงับ"
    flash(f"{status}บัญชี '{user.username}' แล้ว", "success")
    return redirect(url_for("admin.manage_users"))


# ── Imported users (จัดการเป็นชุด) ─────────────────────────

@admin_bp.route("/users/imported")
@admin_required
def imported_users():
    """หน้าจัดการผู้ใช้ role=voter ทั้งหมด (นำเข้า Excel + สมัครเอง) — แก้ไข/ลบเป็นชุด"""
    from models.user import User
    users = User.get_voters()
    return render_template("admin/imported_users.html", users=users)


@admin_bp.route("/users/imported/bulk-delete", methods=["POST"])
@admin_required
def bulk_delete_users():
    """ลบผู้ใช้ที่เลือกหลายคนพร้อมกัน (เฉพาะ role=voter)"""
    from models.user import User
    ids = request.form.getlist("ids", type=int)
    ids = [i for i in ids if i != current_user.id]
    if not ids:
        flash("ยังไม่ได้เลือกผู้ใช้", "warning")
        return redirect(url_for("admin.imported_users"))
    n = User.bulk_delete(ids)
    flash(f"ลบผู้ใช้ {n} บัญชีแล้ว", "info")
    return redirect(url_for("admin.imported_users"))


@admin_bp.route("/users/<int:user_id>/update", methods=["POST"])
@admin_required
def update_user_profile(user_id: int):
    """แก้ไขชื่อ-สกุล / username / email / รหัสผ่าน ของผู้ใช้รายคน"""
    from models.user import User
    user = User.get_by_id_any(user_id)
    if not user:
        abort(404)
    full_name = request.form.get("full_name", "").strip()
    username  = request.form.get("username", "").strip()
    email     = request.form.get("email", "").strip().lower()
    password  = request.form.get("password", "").strip()
    try:
        user.update_profile(
            full_name=full_name or None,
            username=username or None,
            email=email or None,
        )
        if password:
            user.set_password(password)
        flash(f"บันทึกข้อมูลของ '{user.full_name}' แล้ว", "success")
    except Exception:
        flash("แก้ไขไม่สำเร็จ — อาจมีชื่อ-สกุล / username / email ซ้ำกับคนอื่น", "danger")
    return redirect(url_for("admin.imported_users"))


# ── Members import (สร้างเป็น users — ล็อกอินด้วยรหัสผ่านได้) ──

@admin_bp.route("/members/import", methods=["GET", "POST"])
@admin_required
def import_members():
    from models.user import User

    if request.method == "POST":
        f = request.files.get("file")
        if not f or not f.filename.endswith((".xlsx", ".xls")):
            flash("กรุณาอัปโหลดไฟล์ Excel (.xlsx / .xls)", "danger")
            return redirect(url_for("admin.import_members"))
        try:
            import openpyxl
            wb   = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws   = wb.active
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                full_name = str(row[0]).strip()
                username  = str(row[1]).strip()         if len(row) > 1 and row[1] else None
                email     = str(row[2]).strip().lower() if len(row) > 2 and row[2] else None
                password  = str(row[3]).strip()         if len(row) > 3 and row[3] else None
                rows.append({
                    "full_name": full_name,
                    "username":  username,
                    "email":     email,
                    "password":  password,
                })

            result = User.import_from_rows(rows)
            flash(
                f"นำเข้าสำเร็จ — เพิ่มใหม่ {result['added']} คน, "
                f"อัปเดต {result['updated']} คน, "
                f"ข้าม {result['skipped']} คน",
                "success",
            )
        except Exception as e:
            flash(f"เกิดข้อผิดพลาด: {e}", "danger")
        return redirect(url_for("admin.import_members"))

    total = len(User.get_all())
    return render_template("admin/import_members.html", total=total)


# ── System settings ────────────────────────────────────────

@admin_bp.route("/system", methods=["GET", "POST"])
@admin_required
def system():
    if request.method == "POST":
        for key in ("verify_enabled", "vote_enabled"):
            val = "1" if request.form.get(key) else "0"
            SystemSetting.set(key, val)
        flash("บันทึกการตั้งค่าระบบแล้ว", "success")
        return redirect(url_for("admin.system"))

    settings = SystemSetting.get_all()
    return render_template("admin/system.html", settings=settings)


# ── Logs ───────────────────────────────────────────────────

@admin_bp.route("/logs")
@admin_required
def logs():
    keyword     = request.args.get("keyword", "")
    ip          = request.args.get("ip", "")
    date_from   = request.args.get("date_from", "")
    date_to     = request.args.get("date_to", "")
    system_type = request.args.get("system_type", "")

    log_rows = AccessLog.search(
        keyword=keyword, ip=ip,
        date_from=date_from, date_to=date_to,
        system_type=system_type,
    )
    return render_template(
        "admin/logs.html",
        logs=log_rows,
        keyword=keyword, ip=ip,
        date_from=date_from, date_to=date_to,
        system_type=system_type,
    )


# ── Reports ────────────────────────────────────────────────

@admin_bp.route("/reports")
@admin_required
def reports():
    elections       = Election.get_all()
    total_members   = len(Member.get_all())
    total_verified  = Member.count_verified()
    total_changed   = len(Member.get_email_changed())
    return render_template(
        "admin/reports.html",
        elections=elections,
        total_members=total_members,
        total_verified=total_verified,
        total_changed=total_changed,
    )


# ── Reset ──────────────────────────────────────────────────

@admin_bp.route("/reset", methods=["POST"])
@admin_required
def reset():
    """ลบข้อมูลผลเลือกตั้งและการยืนยันตัวตนทั้งหมด"""
    confirm = request.form.get("confirm", "")
    if confirm != "RESET":
        flash("กรุณายืนยันโดยพิมพ์ RESET", "danger")
        return redirect(url_for("admin.dashboard"))
    from db import get_db
    conn = get_db()
    cur  = conn.cursor()
    cur.execute("DELETE FROM votes")
    cur.execute("DELETE FROM vote_turnout")
    cur.execute("DELETE FROM otps")
    cur.execute("UPDATE members SET verified = FALSE, email_new = NULL")
    cur.execute("DELETE FROM access_logs")
    conn.commit()
    cur.close()
    flash("ลบข้อมูลผลเลือกตั้งและการยืนยันตัวตนทั้งหมดแล้ว", "success")
    return redirect(url_for("admin.dashboard"))


# ── 403 handler ────────────────────────────────────────────

@admin_bp.errorhandler(403)
def forbidden(e):
    return render_template("errors/403.html"), 403

# ── Logs export ────────────────────────────────────────────
@admin_bp.route("/logs/export")
@admin_required
def export_logs():
    fmt = request.args.get("format", "excel")
    rows_raw = AccessLog.get_all_raw(limit=5000)
    rows = [
        (r.get("member_name", ""), r.get("action", ""), r.get("ip_address", ""),
         r.get("system_type", ""), str(r.get("logged_at", "")))
        for r in rows_raw
    ]
    return make_table_response(
        fmt,
        sheet_title="Log การใช้งาน",
        headers=["ชื่อสมาชิก", "action", "IP", "ระบบ", "เวลา"],
        rows=rows, col_widths=[25, 20, 18, 12, 22],
        filename_base="logs",
    )


# ── Reports ────────────────────────────────────────────────
@admin_bp.route("/reports/verified/export")
@admin_required
def export_verified():
    fmt = request.args.get("format", "excel")
    members = Member.get_verified()
    rows = [(i + 1, m.full_name, m.email or "", m.email_new or "")
            for i, m in enumerate(members)]
    return make_table_response(
        fmt,
        sheet_title="ยืนยันตัวตนแล้ว",
        headers=["ลำดับ", "ชื่อ-สกุล", "Email เดิม", "Email ใหม่"],
        rows=rows, col_widths=[8, 30, 30, 30],
        filename_base="verified",
    )


@admin_bp.route("/reports/email-changed/export")
@admin_required
def export_email_changed():
    fmt = request.args.get("format", "excel")
    members = Member.get_email_changed()
    rows = [(i + 1, m.full_name, m.email or "", m.email_new or "")
            for i, m in enumerate(members)]
    return make_table_response(
        fmt,
        sheet_title="เปลี่ยน Email",
        headers=["ลำดับ", "ชื่อ-สกุล", "Email เดิม", "Email ใหม่"],
        rows=rows, col_widths=[8, 30, 30, 30],
        filename_base="email_changed",
    )


@admin_bp.route("/reports/votes/export")
@admin_required
def export_votes():
    """รายชื่อผู้มาใช้สิทธิ (ตรวจสอบได้) — ไม่มีข้อมูลว่าเลือกใคร"""
    fmt = request.args.get("format", "excel")
    elections = Election.get_all()
    rows = []
    for e in elections:
        for v in Turnout.list_by_election(e.id):
            name = v.get("full_name") or f"#{v['member_id']}"
            rows.append((e.title, name, str(v["voted_at"])))
    return make_table_response(
        fmt,
        sheet_title="รายชื่อผู้มาใช้สิทธิ",
        headers=["วาระ", "ชื่อ-สกุลผู้มาใช้สิทธิ", "เวลา"],
        rows=rows, col_widths=[30, 30, 22],
        filename_base="turnout",
    )


@admin_bp.route("/reports/summary/export")
@admin_required
def export_summary():
    fmt = request.args.get("format", "excel")
    elections = Election.get_all()
    rows = [(e.title, e.type_label, e.status, Turnout.count_by_election(e.id))
            for e in elections]
    return make_table_response(
        fmt,
        sheet_title="สรุปจำนวนผู้มาลงคะแนน",
        headers=["วาระ", "ประเภท", "สถานะ", "จำนวนผู้ลงคะแนน"],
        rows=rows, col_widths=[30, 20, 12, 18],
        filename_base="summary",
    )


@admin_bp.route("/reports/results/export")
@admin_required
def export_results_all():
    fmt = request.args.get("format", "excel")
    elections = Election.get_all()
    rows = []
    for e in elections:
        candidates = Candidate.get_by_election_with_votes(e.id)
        total = sum(c.vote_count for c in candidates)
        for c in candidates:
            pct = f"{round(c.vote_count / total * 100, 1)}%" if total else "0%"
            rows.append((e.title, c.number, c.name, c.vote_count, pct))
    return make_table_response(
        fmt,
        sheet_title="สรุปผลคะแนน",
        headers=["วาระ", "หมายเลข", "ชื่อ", "คะแนน", "เปอร์เซ็นต์"],
        rows=rows, col_widths=[30, 10, 30, 12, 12],
        filename_base="results_all",
    )
