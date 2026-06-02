"""
routes/candidates.py — Candidates Blueprint

GET  /candidates                   — รายชื่อผู้สมัครทุกวาระ
GET  /candidates/<election_id>     — ผู้สมัครของวาระใดวาระหนึ่ง
GET  /candidates/<election_id>/export — export Excel
"""

import io
from datetime import datetime

from flask import Blueprint, render_template, redirect, url_for, flash, send_file
from flask_login import login_required

from models.election  import Election
from models.candidate import Candidate

candidates_bp = Blueprint("candidates", __name__)


@candidates_bp.route("/candidates")
def list_all():
    """หน้ารวมผู้สมัครทุกวาระ — จัดกลุ่มตามวาระ"""
    elections  = Election.get_all()
    by_election = {
        e: Candidate.get_by_election(e.id) for e in elections
    }
    return render_template("candidates.html", by_election=by_election)


@candidates_bp.route("/candidates/<int:election_id>")
def by_election(election_id: int):
    election = Election.get_by_id(election_id)
    if not election:
        flash("ไม่พบวาระการเลือกตั้ง", "danger")
        return redirect(url_for("candidates.list_all"))

    candidates = Candidate.get_by_election(election_id)
    return render_template(
        "candidates.html",
        by_election={election: candidates},
        single=True,
    )


@candidates_bp.route("/candidates/<int:election_id>/export")
@login_required
def export_excel(election_id: int):
    """Export รายชื่อผู้สมัครเป็น Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    election = Election.get_by_id(election_id)
    if not election:
        flash("ไม่พบวาระการเลือกตั้ง", "danger")
        return redirect(url_for("candidates.list_all"))

    candidates = Candidate.get_by_election_with_votes(election_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ผู้สมัคร"

    # Header
    headers = ["หมายเลข", "ชื่อ", "พรรค/กลุ่ม", "นโยบาย/ประวัติ", "คะแนนที่ได้"]
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, c in enumerate(candidates, 2):
        ws.cell(row=row_idx, column=1, value=c.number)
        ws.cell(row=row_idx, column=2, value=c.name)
        ws.cell(row=row_idx, column=3, value=c.party or "")
        ws.cell(row=row_idx, column=4, value=c.bio   or "")
        ws.cell(row=row_idx, column=5, value=c.vote_count)

    # Column widths
    for col, width in zip("ABCDE", [10, 30, 25, 50, 15]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"candidates_{election_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
