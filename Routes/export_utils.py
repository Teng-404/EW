"""
routes/export_utils.py — ตัวช่วยสร้างไฟล์รายงาน (Excel + PDF)

รวมโค้ดสร้างไฟล์ export ไว้ที่เดียว เพื่อไม่ให้ซ้ำกันหลายไฟล์
ใช้งานหลักผ่าน make_table_response():
    return make_table_response(
        fmt,                       # 'pdf' → PDF, อื่น ๆ (เช่น 'excel') → xlsx
        sheet_title="...",         # ชื่อชีต / หัวรายงาน
        headers=[...],             # หัวคอลัมน์
        rows=[(...), (...)],       # ข้อมูล
        col_widths=[...],          # ความกว้างคอลัมน์ (หน่วยเดียวกับ openpyxl)
        filename_base="verified",  # ชื่อไฟล์ (จะต่อท้ายด้วยวันที่ + นามสกุล)
    )

หมายเหตุภาษาไทยใน PDF:
    reportlab ใช้ฟอนต์ default (Helvetica) ที่แสดงไทยไม่ได้
    โมดูลนี้จะมองหาไฟล์ฟอนต์ไทย (.ttf) ตามลำดับใน _register_thai_font()
    ต้องมีฟอนต์ไทยอยู่ในระบบ มิฉะนั้น PDF จะ fallback เป็น Helvetica
    (ตัวอักษรไทยจะหาย) — ดูวิธีติดตั้งฟอนต์ใน README/คอมเมนต์ด้านล่าง
"""

import io
import os
from datetime import datetime

from flask import send_file, current_app

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ══════════════════════════════════════════════════════════
#  Excel
# ══════════════════════════════════════════════════════════

def make_xlsx(ws_title: str, headers: list, rows: list, col_widths: list) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ws_title

    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════
#  PDF
# ══════════════════════════════════════════════════════════

_THAI_REGULAR = "ThaiSans"
_THAI_BOLD    = "ThaiSans-Bold"
_thai_state   = None   # None=ยังไม่ลอง, True=ลงทะเบียนสำเร็จ, False=ไม่พบฟอนต์


def _register_thai_font() -> bool:
    """
    ลงทะเบียนฟอนต์ไทยให้ reportlab (ทำครั้งเดียวต่อ process)
    คืน True ถ้าพบและลงทะเบียนได้ / False ถ้าไม่พบฟอนต์ไทยเลย
    """
    global _thai_state
    if _thai_state is not None:
        return _thai_state

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    try:
        static = current_app.static_folder
    except Exception:
        static = None

    # ── ที่ ๆ จะมองหาฟอนต์ (เรียงตามลำดับความสำคัญ) ──
    regular_paths, bold_paths = [], []
    if static:
        fonts_dir = os.path.join(static, "fonts")
        regular_paths += [
            os.path.join(fonts_dir, "Sarabun-Regular.ttf"),
            os.path.join(fonts_dir, "THSarabunNew.ttf"),
            os.path.join(fonts_dir, "Sarabun.ttf"),
        ]
        bold_paths += [
            os.path.join(fonts_dir, "Sarabun-Bold.ttf"),
            os.path.join(fonts_dir, "THSarabunNew-Bold.ttf"),
        ]
    # system fonts (เช่น apt-get install fonts-tlwg-sarabun)
    regular_paths += [
        "/usr/share/fonts/truetype/tlwg/Sarabun.ttf",
        "/usr/share/fonts/truetype/thai-tlwg/Sarabun.ttf",
        "/usr/share/fonts/truetype/tlwg/Loma.ttf",
    ]
    bold_paths += [
        "/usr/share/fonts/truetype/tlwg/Sarabun-Bold.ttf",
        "/usr/share/fonts/truetype/tlwg/Loma-Bold.ttf",
    ]

    regular = next((p for p in regular_paths if os.path.exists(p)), None)
    if not regular:
        _thai_state = False
        try:
            current_app.logger.warning(
                "export_utils: ไม่พบฟอนต์ไทย — PDF ภาษาไทยจะแสดงไม่ได้ "
                "(วางไฟล์ .ttf ไว้ที่ static/fonts/ หรือ apt-get install fonts-tlwg-sarabun)"
            )
        except Exception:
            pass
        return False

    pdfmetrics.registerFont(TTFont(_THAI_REGULAR, regular))

    bold = next((p for p in bold_paths if os.path.exists(p)), None)
    pdfmetrics.registerFont(TTFont(_THAI_BOLD, bold or regular))

    _thai_state = True
    return True


def _esc(s) -> str:
    """escape อักขระพิเศษก่อนใส่ใน Paragraph (reportlab parse mini-HTML)"""
    text = "" if s is None else str(s)
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def make_pdf(title: str, headers: list, rows: list, col_widths: list = None) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    has_thai  = _register_thai_font()
    base_font = _THAI_REGULAR if has_thai else "Helvetica"
    bold_font = _THAI_BOLD    if has_thai else "Helvetica-Bold"

    # คอลัมน์เยอะ → ใช้แนวนอนกันตาราง overflow
    pagesize = landscape(A4) if len(headers) >= 5 else A4

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=16 * mm, bottomMargin=14 * mm,
        title=title,
    )

    # แปลงความกว้างคอลัมน์ให้พอดีหน้า (สัดส่วนเดิม)
    if col_widths and len(col_widths) == len(headers):
        total   = sum(col_widths) or 1
        widths  = [doc.width * w / total for w in col_widths]
    else:
        widths  = [doc.width / len(headers)] * len(headers)

    head_style = ParagraphStyle("h", fontName=bold_font, fontSize=9.5,
                                leading=12, textColor=colors.white, alignment=TA_CENTER)
    cell_style = ParagraphStyle("c", fontName=base_font, fontSize=9, leading=12)

    data = [[Paragraph(_esc(h), head_style) for h in headers]]
    for row in rows:
        data.append([Paragraph(_esc(v), cell_style) for v in row])

    table = Table(data, colWidths=widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",     (0, 0), (-1, 0), colors.HexColor("#1F3A5F")),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("GRID",           (0, 0), (-1, -1), 0.5, colors.HexColor("#D7DEE8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FD")]),
        ("TOPPADDING",     (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
        ("LEFTPADDING",    (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
    ]))

    title_style = ParagraphStyle("t", fontName=bold_font, fontSize=15,
                                 leading=20, textColor=colors.HexColor("#0D1F3C"))
    meta_style  = ParagraphStyle("m", fontName=base_font, fontSize=8,
                                 textColor=colors.HexColor("#8A96A3"))
    stamp = datetime.now().strftime("%d/%m/%Y %H:%M")

    story = [
        Paragraph(_esc(title), title_style),
        Spacer(1, 3 * mm),
        Paragraph(_esc(f"ออกรายงานเมื่อ {stamp} — รวม {len(rows)} รายการ"), meta_style),
        Spacer(1, 5 * mm),
        table,
    ]
    doc.build(story)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════
#  Dispatcher — เลือก format
# ══════════════════════════════════════════════════════════

def make_table_response(fmt: str, *, sheet_title: str, headers: list,
                        rows: list, col_widths: list, filename_base: str):
    """สร้าง response ไฟล์ตาม fmt ('pdf' → PDF, อื่น ๆ → xlsx)"""
    stamp = datetime.now().strftime("%Y%m%d")
    if (fmt or "").lower() == "pdf":
        buf = make_pdf(sheet_title, headers, rows, col_widths)
        return send_file(buf, mimetype="application/pdf",
                         as_attachment=True,
                         download_name=f"{filename_base}_{stamp}.pdf")

    buf = make_xlsx(sheet_title, headers, rows, col_widths)
    return send_file(buf, mimetype=XLSX_MIME,
                     as_attachment=True,
                     download_name=f"{filename_base}_{stamp}.xlsx")
